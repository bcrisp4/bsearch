"""Behaviour of the docx/xlsx builders (data-driven, deterministic output)."""

from __future__ import annotations

import io
from typing import TYPE_CHECKING

from docx import Document
from openpyxl import load_workbook

from corpusgen.office import build_docx_letter, build_xlsx_sheet

if TYPE_CHECKING:
    from pathlib import Path

LETTER = {
    "id": "test-letter",
    "sender_lines": ["Ferrand Bank", "PO Box 1"],
    "date": "4 August 2025",
    "recipient_lines": ["Jane Doe", "1 Test Street"],
    "subject": "Confirmation of address",
    "paragraphs": [
        "Dear Jane Doe,",
        "We confirm the address we hold for you.",
        "Yours sincerely,",
    ],
    "signature_lines": ["A Clerk", "Customer Services"],
}

SHEET = {
    "id": "test-sheet",
    "sheet_name": "Interest 2025",
    "rows": [
        ["Account", "Interest"],
        ["Savings", "47.21"],
        ["Total", "47.21"],
    ],
}


class TestDocxLetter:
    def test_contains_content(self, tmp_path: Path) -> None:
        out = tmp_path / "letter.docx"

        build_docx_letter(LETTER, out)

        text = "\n".join(p.text for p in Document(str(out)).paragraphs)
        assert "Confirmation of address" in text
        assert "Jane Doe" in text

    def test_rerun_is_byte_identical(self, tmp_path: Path) -> None:
        a, b = tmp_path / "a.docx", tmp_path / "b.docx"

        build_docx_letter(LETTER, a)
        build_docx_letter(LETTER, b)

        assert a.read_bytes() == b.read_bytes()


class TestXlsxSheet:
    def test_contains_rows(self, tmp_path: Path) -> None:
        out = tmp_path / "sheet.xlsx"

        build_xlsx_sheet(SHEET, out)

        wb = load_workbook(io.BytesIO(out.read_bytes()))
        ws = wb["Interest 2025"]
        assert ws["A1"].value == "Account"
        assert ws["B2"].value == "47.21"

    def test_rerun_is_byte_identical(self, tmp_path: Path) -> None:
        a, b = tmp_path / "a.xlsx", tmp_path / "b.xlsx"

        build_xlsx_sheet(SHEET, a)
        build_xlsx_sheet(SHEET, b)

        assert a.read_bytes() == b.read_bytes()
